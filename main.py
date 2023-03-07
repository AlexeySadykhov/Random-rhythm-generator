import functions as func
import generators as gen
import sys
import openpyxl as pyxl
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QMessageBox, QFileDialog, QInputDialog
from main_window_design import Ui_TabWidget


class MainWindow(QtWidgets.QTabWidget):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.ui = Ui_TabWidget()
        self.ui.setupUi(self)
        self.ui.rr_generate_button.clicked.connect(self.rr_generate)
        self.ui.ir_get_from_excel_button.clicked.connect(self.get_from_excel)
        self.ui.ir_generate_button.clicked.connect(self.ir_generate)
        self.ui.pr_generate_button.clicked.connect(self.pr_generate)
        self.ui.pe_generate_button.clicked.connect(self.pe_generate)

    def save_file(self, data):
        QMessageBox.information(self, "Ready", "Rhythm has been generated.",
                                QMessageBox.Ok)
        out_file_opt = QFileDialog.Options()
        out_file_name, done = QFileDialog.getSaveFileName(self, "Save file", "",
                                                          "Text files (*.txt)",
                                                          options=out_file_opt)
        if not done:
            return
        out_file = open(out_file_name, "w")
        out_file.write(data)
        out_file.close()

    def show_error(self, text):
        QMessageBox.critical(self, "Error", text,
                             QMessageBox.Ok)

    def rr_generate(self):
        values = [self.ui.rr_time_sig_first_spinBox.value(),
                  self.ui.rr_time_sig_second_spinBox.value(),
                  self.ui.rr_num_of_bars_spinBox.value(),
                  self.ui.rr_num_of_staves_spinBox.value()]
        if not func.check_values(values):
            self.show_error("Specify all arguments.")
        elif self.ui.rr_den_min_spinBox.value() > self.ui.rr_den_max_spinBox.value():
            self.show_error("Min denominator must be lower then max denominator.")
        else:
            data = gen.generate_regular_rhythm(self.ui.rr_main_dur_comboBox.currentIndex() + 1,
                                               [self.ui.rr_time_sig_first_spinBox.value(),
                                                self.ui.rr_time_sig_second_spinBox.value()],
                                               self.ui.rr_num_of_bars_spinBox.value(),
                                               self.ui.rr_num_of_staves_spinBox.value(),
                                               self.ui.rr_rests_perc_spinBox.value(),
                                               self.ui.rr_div_perc_spinBox.value(),
                                               self.ui.rr_den_min_spinBox.value(),
                                               self.ui.rr_den_max_spinBox.value(),
                                               self.ui.rr_div_parts_spinBox.value(),
                                               self.ui.rr_division_checkBox.isChecked(),
                                               self.ui.rr_div_irreg_checkBox.isChecked(),
                                               self.ui.rr_rests_checkBox.isChecked())
            self.save_file(data)

    def get_from_excel(self):
        number_of_layers = self.ui.ir_num_of_layers_spinBox.value()
        number_of_staves = self.ui.ir_num_of_staves_spinBox.value()
        input_vals = [self.ui.ir_num_of_staves_spinBox.value(),
                      self.ui.ir_time_sig_den_spinBox.value()]
        if not func.check_values(input_vals):
            self.show_error("Specify all arguments.")
        else:
            in_file_opt = QFileDialog.Options()
            in_file_location, done = QFileDialog.getOpenFileName(self, "Open file", "",
                                                                 "Excel files (*.xlsx)",
                                                                 options=in_file_opt)
            if not done:
                return
            in_workbook = pyxl.load_workbook(in_file_location)
            worksheet = None
            if len(in_workbook.sheetnames) > 1:
                sheet_name, done = QInputDialog.getText(self, "Warning", """Your file has several worksheets. 
Enter the name of sheet you want to parse:""")
                if done and sheet_name:
                    try:
                        worksheet = in_workbook[sheet_name]
                    except KeyError:
                        self.show_error(f"Error. There is no {sheet_name} sheet.")
                        return
                if done and not sheet_name:
                    self.show_error("Error. No sheet name given.")
                    return
                if not done:
                    return
            else:
                worksheet = in_workbook.active

            values = []
            for i in range(1, worksheet.max_row + 1):
                row = []
                for j in range(1, worksheet.max_column + 1):
                    val = worksheet.cell(row=i, column=j).value
                    if type(val) == str:
                        self.show_error("Error. All values must be digits.")
                        return
                    else:
                        row.append(val)
                values.append(row)

            values = list(filter(func.item_exists, values))
            if len(values) != number_of_layers * number_of_staves:
                self.show_error(f"""Error. Number of layers must be {number_of_layers} 
and number of staves must be {number_of_staves}.""")
                return

            values = func.group(values, number_of_layers)
            for staff in values:
                rows_lengths = [len(func.list_filter(row)) for row in staff]
                if not func.is_sorted_desc(rows_lengths) or func.is_equal(rows_lengths):
                    self.show_error("Error. Input is incorrect. Not a tree.")
                    return
            data = gen.create_rhythm_from_excel(values, self.ui.ir_time_sig_den_spinBox.value())
            self.save_file(data)

    def ir_generate(self):
        values = [self.ui.ir_num_of_staves_spinBox.value(),
                  self.ui.ir_num_of_bars_spinBox.value(),
                  self.ui.ir_min_val_spinBox.value(),
                  self.ui.ir_max_val_spinBox.value()]
        if self.ui.ir_single_time_sig_checkBox.isChecked():
            values.extend([self.ui.ir_time_sig_first_spinBox.value(),
                           self.ui.ir_time_sig_second_spinBox.value()])
        else:
            values.append(self.ui.ir_time_sig_den_spinBox.value())
        if not func.check_values(values):
            self.show_error("Specify all arguments.")
        elif self.ui.ir_min_val_spinBox.value() >= self.ui.ir_max_val_spinBox.value():
            self.show_error("Min value must be lower then max value.")
        else:
            data = gen.generate_irregular_rhythm(self.ui.ir_min_val_spinBox.value(),
                                                 self.ui.ir_max_val_spinBox.value(),
                                                 self.ui.ir_num_of_layers_spinBox.value(),
                                                 self.ui.ir_num_of_bars_spinBox.value(),
                                                 self.ui.ir_num_of_staves_spinBox.value(),
                                                 self.ui.ir_div_last_layer_checkBox.isChecked(),
                                                 self.ui.ir_rests_perc_spinBox.value(),
                                                 self.ui.ir_rests_checkBox.isChecked(),
                                                 self.ui.ir_single_time_sig_checkBox.isChecked(),
                                                 [self.ui.ir_time_sig_first_spinBox.value(),
                                                  self.ui.ir_time_sig_second_spinBox.value()],
                                                 self.ui.ir_time_sig_den_spinBox.value())
            self.save_file(data)

    def pr_generate(self):
        values = [self.ui.pr_min_phase_spinBox.value(),
                  self.ui.pr_max_phase_spinBox.value()]
        if self.ui.pr_gen_rand_ph_radioButton.isChecked():
            values.append(self.ui.pr_num_of_staves_spinBox.value())
            if self.ui.pr_diff_time_sig_checkBox.isChecked():
                values.append(self.ui.pr_time_sig_lst_lineEdit.text())
                if not func.check_values(values):
                    self.show_error("Specify all arguments.")
                elif self.ui.pr_min_phase_spinBox.value() > self.ui.pr_max_phase_spinBox.value():
                    self.show_error("Min phase must be lower then max phase.")
                elif not func.check_expression(self.ui.pr_time_sig_lst_lineEdit.text(), mode="time_sig"):
                    self.show_error("Time signature list values are not correct.")
                else:
                    data = gen.generate_random_phases(self.ui.pr_num_of_staves_spinBox.value(),
                                                      self.ui.pr_time_sig_lst_lineEdit.text(),
                                                      self.ui.pr_num_of_bars_spinBox.value(),
                                                      self.ui.pr_min_phase_spinBox.value(),
                                                      self.ui.pr_max_phase_spinBox.value(),
                                                      self.ui.pr_rests_perc_spinBox.value(),
                                                      self.ui.pr_rests_checkBox.isChecked())
                    self.save_file(data)
            else:
                values.extend([self.ui.pr_time_sig_first_spinBox.value(),
                               self.ui.pr_time_sig_second_spinBox.value(),
                               self.ui.pr_num_of_bars_spinBox.value()])
                if not func.check_values(values):
                    self.show_error("Specify all arguments.")
                elif self.ui.pr_min_phase_spinBox.value() > self.ui.pr_max_phase_spinBox.value():
                    self.show_error("Min phase must be lower then max phase.")
                else:
                    data = gen.generate_random_phases(self.ui.pr_num_of_staves_spinBox.value(),
                                                      [[self.ui.pr_time_sig_first_spinBox.value(),
                                                        self.ui.pr_time_sig_second_spinBox.value()]],
                                                      self.ui.pr_num_of_bars_spinBox.value(),
                                                      self.ui.pr_min_phase_spinBox.value(),
                                                      self.ui.pr_max_phase_spinBox.value(),
                                                      self.ui.pr_rests_perc_spinBox.value(),
                                                      self.ui.pr_rests_checkBox.isChecked())
                    self.save_file(data)
        else:
            values.extend([self.ui.pr_time_sig_first_spinBox.value(),
                           self.ui.pr_time_sig_second_spinBox.value(),
                           self.ui.pr_num_of_bars_spinBox.value(),
                           self.ui.pr_num_of_imp_spinBox.value()])
            if not func.check_values(values):
                self.show_error("Specify all arguments.")
            elif self.ui.pr_min_phase_spinBox.value() > self.ui.pr_max_phase_spinBox.value():
                self.show_error("Min phase must be lower then max phase.")
            else:
                data = gen.place_points_on_phases([self.ui.pr_time_sig_first_spinBox.value(),
                                                   self.ui.pr_time_sig_second_spinBox.value()],
                                                  self.ui.pr_num_of_bars_spinBox.value(),
                                                  self.ui.pr_min_phase_spinBox.value(),
                                                  self.ui.pr_max_phase_spinBox.value(),
                                                  self.ui.pr_num_of_imp_spinBox.value(),
                                                  self.ui.pr_phase_filter_checkBox.isChecked())
                if not data:
                    self.show_error("It is impossible to place these values.")
                else:
                    self.save_file(data)

    def pe_generate(self):
        values = [self.ui.pe_time_sig_first_spinBox.value(),
                  self.ui.pe_time_sig_second_spinBox.value(),
                  self.ui.pe_num_of_bars_spinBox.value(),
                  self.ui.pe_num_of_staves_spinBox.value(),
                  self.ui.pe_durs_of_ev_lineEdit.text()]
        if not func.check_values(values):
            self.show_error("Specify all arguments.")
        elif not func.check_expression(self.ui.pe_durs_of_ev_lineEdit.text()):
            self.show_error("Duration list values are not correct. Enter them splitting by space.")
        else:
            data = gen.place_events(self.ui.pe_main_dur_comboBox.currentIndex() + 1,
                                    [self.ui.pe_time_sig_first_spinBox.value(),
                                     self.ui.pe_time_sig_second_spinBox.value()],
                                    self.ui.pe_num_of_bars_spinBox.value(),
                                    self.ui.pe_num_of_staves_spinBox.value(),
                                    self.ui.pe_durs_of_ev_lineEdit.text())
            if not data:
                self.show_error("It is impossible to place these values.")
            else:
                self.save_file(data)


app = QtWidgets.QApplication(sys.argv)
application = MainWindow()
application.show()

sys.exit(app.exec_())
